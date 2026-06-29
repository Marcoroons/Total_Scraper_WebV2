"""
ecom_export.py — Build the Excel workbook for the Competitor Analysis exporter.

Layer A: Bahasa Indonesia parser
   - bundle / pack count   (isi N, N pcs, Nx, lusin, ...)
   - unit volume + UOM     (Nml / NL → ml; Ng / Nkg → g)
   - container type        (kaleng / kotak / botol / pouch ...)
   - flavour keywords      (coklat, stroberi, ayam bawang, ...)
   - reviews count         (best-effort from raw_payload across actor field shapes)

Layer B: aggregation
   - Group by brand_name (primary) and by (brand_name × flavour) (secondary)
   - Aggregate: median per-unit cost, sum sold, sum reviews, avg rating,
                typical bundle volume (median unit_volume × median total_units)
   - Sort by total_sold desc — "most popular first"

Layer C: openpyxl workbook
   Sheets:
     "Products"     — one row per brand, flavours collected (the user-spec view)
     "By Flavour"   — one row per (brand × flavour)
     "Raw Listings" — every parsed row for drill-down
     "Notes"        — parsing caveats + column definitions

This file runs entirely in the export-service. Nothing is persisted back to
ecom_listings — the parser runs fresh on every export so we can iterate the
regex without backfill jobs. Phase 2 (persist) can come later.
"""
from __future__ import annotations

import io
import re
import statistics as stats
import unicodedata
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _norm_text(s: str) -> str:
    """Lowercase + strip diacritics. 'Nestlé' → 'nestle'. Matches the worker's
    normalization so shop-name matches behave identically on both sides."""
    return "".join(
        c for c in unicodedata.normalize("NFKD", s or "") if not unicodedata.combining(c)
    ).lower()


def _shop_tokens(s: str) -> list:
    """Token split after normalization — used for the specific_shops match."""
    return [t for t in re.findall(r"[a-z0-9]+", _norm_text(s)) if t]


def _apply_export_shop_filter(rows: list[dict], shop_filter: str | None,
                              specific_shops: list[str] | None) -> list[dict]:
    """Filter the listings list by the export-time shop filter.
    Matches the worker's filter semantics (see _shop_is_official /
    _apply_official_filter in worker.py)."""
    mode = (shop_filter or "all").lower()
    if mode == "all" or not mode:
        return rows
    if mode == "official_only":
        return [r for r in rows if r.get("is_official_store")]
    if mode == "non_official_only":
        return [r for r in rows if not r.get("is_official_store")]
    if mode == "specific_shops":
        groups = []
        for s in (specific_shops or []):
            toks = _shop_tokens(s)
            if toks:
                groups.append(toks)
        if not groups:
            return []
        return [
            r for r in rows
            if any(all(t in _norm_text(r.get("shop_name") or "") for t in g) for g in groups)
        ]
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# LAYER A — Bahasa parser
# ═════════════════════════════════════════════════════════════════════════════

# Flavour vocabulary. Lowercased exact-substring match against the title; the
# first hit wins so longer / more specific names go first.
FLAVOUR_KEYWORDS: list[str] = [
    # Indomie / instant noodle flavours
    "ayam bawang", "kari ayam", "soto mie", "soto", "mi goreng", "mie goreng",
    "rendang", "kaldu ayam", "cabe ijo", "sambal matah", "salted egg",
    # Sweet / dairy
    "coklat", "cokelat", "chocolate", "vanilla", "vanila", "stroberi", "strawberry",
    "mangga", "mango", "pisang", "banana", "melon", "taro", "matcha",
    "kopi", "coffee", "latte", "cappuccino", "mocca", "mocha",
    # Tea / others
    "teh tarik", "lychee", "leci", "anggur", "jeruk", "lemon",
    # Generic
    "original", "plain", "tawar", "pedas", "manis", "asin", "gurih",
]

# Container terms (lowercased substring match).
CONTAINER_KEYWORDS: dict[str, list[str]] = {
    "can":    ["kaleng", "can "],
    "box":    ["kotak", "karton", "dus", " box"],
    "bottle": ["botol", "bottle"],
    "pouch":  ["pouch", "sachet", "renceng", "saset"],
}

# Volume regexes — capture the number + the (lowercased) unit token.
_VOLUME_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(\d+(?:[.,]\d+)?)\s*ml\b",        re.I), "ml"),
    (re.compile(r"(\d+(?:[.,]\d+)?)\s*l(?:iter)?\b", re.I), "L"),
    (re.compile(r"(\d+(?:[.,]\d+)?)\s*g(?:r|ram)?\b", re.I), "g"),
    (re.compile(r"(\d+(?:[.,]\d+)?)\s*kg\b",        re.I), "kg"),
]

# Pack-count regexes. Order matters — earlier patterns win.
_PACK_PATTERNS: list[tuple[re.Pattern, str | None]] = [
    (re.compile(r"isi\s*(\d+)",                          re.I), None),
    (re.compile(r"(\d+)\s*(?:pcs|pieces|pack|pak|sachet|saset)\b", re.I), None),
    (re.compile(r"\b(\d+)\s*x\b",                        re.I), None),
    (re.compile(r"\bx\s*(\d+)\b",                        re.I), None),
    (re.compile(r"(\d+)\s*@",                            re.I), None),
    (re.compile(r"\b(\d+)\s*lusin\b",                    re.I), "lusin"),
    (re.compile(r"\b1\s*lusin\b",                        re.I), "single_lusin"),
    (re.compile(r"renceng\s*(\d+)",                      re.I), None),
    (re.compile(r"(\d+)\s*renceng",                      re.I), None),
]


def _norm_num(s: str) -> float:
    """'1.500,5' → 1500.5; '2,5' → 2.5; '1.5' → 1.5. Tolerates Indonesian
    thousands-dot when followed by exactly 3 digits."""
    s = s.strip()
    if "," in s and "." in s:
        # Both — Indonesian convention: dot = thousands, comma = decimal.
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif "." in s:
        # Single dot — thousands if exactly 3 digits after.
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
            s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_pack_count(text: str) -> tuple[int, str]:
    """Return (total_units, confidence). Default to (1, 'no_signal')."""
    if not text:
        return 1, "no_signal"
    t = text.lower()
    for rx, special in _PACK_PATTERNS:
        m = rx.search(t)
        if m:
            if special == "single_lusin":
                return 12, "lusin"
            n = int(m.group(1))
            if special == "lusin":
                return max(1, n) * 12, "lusin_n"
            if 1 <= n <= 999:
                return n, "explicit"
    return 1, "no_signal"


def parse_unit_volume(text: str) -> tuple[float | None, str | None]:
    """Return (volume, uom_normalised) or (None, None).
    Liquids return ('ml'), solids return ('g'). L → ×1000 → ml; kg → ×1000 → g."""
    if not text:
        return None, None
    t = text.lower()
    for rx, uom_raw in _VOLUME_PATTERNS:
        m = rx.search(t)
        if m:
            val = _norm_num(m.group(1))
            if val <= 0:
                continue
            if uom_raw == "ml":  return val, "ml"
            if uom_raw == "L":   return val * 1000, "ml"
            if uom_raw == "g":   return val, "g"
            if uom_raw == "kg":  return val * 1000, "g"
    return None, None


def parse_container(text: str) -> str:
    if not text:
        return "other"
    t = text.lower()
    for label, terms in CONTAINER_KEYWORDS.items():
        for term in terms:
            if term in t:
                return label
    return "other"


def parse_flavour(text: str) -> str | None:
    if not text:
        return None
    t = text.lower()
    for kw in FLAVOUR_KEYWORDS:
        if kw in t:
            return kw
    return None


def extract_reviews_count(raw_payload: Any) -> int | None:
    """Best-effort reviews count from the actor's raw response. The two actors
    (gio21/shopee, jupri/tokopedia) use different field names; try the common
    keys and fall back to None so the cell shows '—' rather than a wrong value."""
    if not isinstance(raw_payload, dict):
        return None
    for k in (
        "reviewCount", "review_count", "numReviews", "num_reviews",
        "cmt_count", "ratingCount", "rating_count", "rating_star_count",
        "totalRating", "total_rating", "totalReview", "total_review",
    ):
        v = raw_payload.get(k)
        if v is None:
            continue
        try:
            n = int(float(v))
            if n >= 0:
                return n
        except (TypeError, ValueError):
            continue
    return None


def parse_listing(row: dict) -> dict:
    """Decorate a raw ecom_listings row with parsed fields. Doesn't write back
    to the DB — this is the transient enrichment for one export. Prefers the
    DB columns when the worker already populated them (new product-based jobs
    set `flavour` at scrape time); falls back to regex for legacy rows."""
    title = (row.get("title") or "")
    desc  = (row.get("description") or "")
    blob  = f"{title}  |  {desc}"

    units, _bundle_conf = parse_pack_count(blob)
    vol, uom            = parse_unit_volume(blob)
    container           = parse_container(blob)
    # Worker tag wins; regex is only a fallback for listings ingested before
    # the product-based scraper landed.
    flavour             = (row.get("flavour") or "").strip().lower() or parse_flavour(blob)
    reviews             = extract_reviews_count(row.get("raw_payload"))

    listing_price = row.get("listing_price_idr")
    try:
        listing_price = float(listing_price) if listing_price is not None else None
    except (TypeError, ValueError):
        listing_price = None
    per_unit = (listing_price / units) if (listing_price and units and units > 0) else None

    total_vol = (vol * units) if (vol is not None and units) else None

    sold = row.get("sold_count")
    try:
        sold = int(sold) if sold is not None else None
    except (TypeError, ValueError):
        sold = None

    rating = row.get("rating")
    try:
        rating = float(rating) if rating is not None else None
    except (TypeError, ValueError):
        rating = None

    return {
        **row,
        "_total_units":       units,
        "_unit_volume":       vol,
        "_unit_volume_uom":   uom,
        "_total_volume":      total_vol,
        "_container":         container,
        "_flavour":           flavour,
        "_per_unit_price":    per_unit,
        "_reviews":           reviews,
        "_sold":              sold,
        "_rating":            rating,
        "_listing_price":     listing_price,
    }


# ═════════════════════════════════════════════════════════════════════════════
# LAYER B — aggregation
# ═════════════════════════════════════════════════════════════════════════════

def _median_or_none(xs: list) -> float | None:
    xs = [x for x in xs if x is not None]
    if not xs:
        return None
    return float(stats.median(xs))


def _mean_or_none(xs: list) -> float | None:
    xs = [x for x in xs if x is not None]
    if not xs:
        return None
    return float(stats.mean(xs))


def _sum_or_none(xs: list) -> int | None:
    xs = [x for x in xs if x is not None]
    if not xs:
        return None
    return int(sum(xs))


def _dominant_uom(rows: list[dict]) -> str | None:
    """Pick the most common volume UOM in a group — needed because you can't
    sum ml and g, so we report the bundle volume in the group's majority UOM
    and exclude the other rows from the volume aggregate."""
    counts: dict[str, int] = defaultdict(int)
    for r in rows:
        u = r.get("_unit_volume_uom")
        if u:
            counts[u] += 1
    if not counts:
        return None
    return max(counts.items(), key=lambda x: x[1])[0]


def _price_per_100(row: dict) -> float | None:
    """IDR per 100ml (liquids) / per 100g (solids), per single listing."""
    p   = row.get("_per_unit_price")
    v   = row.get("_unit_volume")
    if p is None or v is None or v <= 0:
        return None
    return float(p) / float(v) * 100.0


def _fmt_user_volume(val, uom) -> str:
    """Format a persisted user-specified volume back into a label like '240 ml'
    or '1 kg'. Used to print the product label in the Products sheet."""
    if val is None or uom is None: return ""
    v = float(val)
    if uom == "ml" and v >= 1000 and (v % 1000) == 0:  return f"{int(v//1000)} L"
    if uom == "g"  and v >= 1000 and (v % 1000) == 0:  return f"{int(v//1000)} kg"
    return f"{int(v) if v.is_integer() else v} {uom}"


def aggregate_by_product(parsed: list[dict]) -> list[dict]:
    """One row per (brand, flavour, container_type, unit_volume, uom) — the
    user's full product tuple. Searches with different specified volumes /
    container types produce separate rows so 'Nescafe Latte 240ml kaleng' and
    'Nescafe Latte 220ml kaleng' don't get squashed together.

    Sales Volume = sum sold; Unit Price per 100ml/g = median of (per_unit_price
    / unit_volume * 100); Top Products = top 3 listings by sold; Reviews = sum.
    Groups with a different UOM are excluded from the per-100 aggregate so
    liquids and solids aren't mixed together."""
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for r in parsed:
        brand   = (r.get("brand_name") or "(unbranded)").strip() or "(unbranded)"
        flavour = (r.get("_flavour") or "").strip() or "(no flavour)"
        # Group key uses the DB columns (which carry the user's intent), NOT
        # the regex-parsed values. When the user didn't specify volume/type,
        # the DB columns are NULL and all such listings cluster into one row.
        ctype       = (r.get("container_type") or "") or None
        db_vol      = r.get("unit_volume")
        db_uom      = r.get("unit_volume_uom")
        try:
            db_vol_f = float(db_vol) if db_vol is not None else None
        except (TypeError, ValueError):
            db_vol_f = None
        key = (brand, flavour, ctype, db_vol_f, db_uom)
        groups[key].append(r)

    out: list[dict] = []
    for (brand, flavour, ctype, db_vol, db_uom), rows in groups.items():
        # Per-100 calculation falls back to parsed unit_volume when the DB
        # column is unset (legacy listings + "no volume specified" products).
        uom_for_per100 = db_uom or _dominant_uom(rows)
        vol_rows = [r for r in rows if r.get("_unit_volume_uom") == uom_for_per100]
        per_100s = [_price_per_100(r) for r in vol_rows]

        # Top 3 listings within this product, ranked by sold count desc.
        ranked = sorted(
            (r for r in rows if r.get("_sold") is not None),
            key=lambda r: -(r["_sold"] or 0),
        )[:3]
        top_products = [
            {
                "title": r.get("title") or "",
                "sold":  r.get("_sold") or 0,
                "url":   r.get("url"),
                "shop":  r.get("shop_name"),
            }
            for r in ranked
        ]

        # Build the product label, adding optional volume + type when present.
        label_bits = [brand]
        if flavour != "(no flavour)":
            label_bits.append(flavour)
        vol_label = _fmt_user_volume(db_vol, db_uom)
        if vol_label:
            label_bits.append(vol_label)
        if ctype:
            label_bits.append(ctype)
        product_label = " ".join(label_bits).strip()

        out.append({
            "brand":             brand,
            "flavour":           flavour,
            "container_type":    ctype,
            "user_volume":       db_vol,
            "user_volume_uom":   db_uom,
            "product_label":     product_label,
            "n_listings":        len(rows),
            "sales_volume":      _sum_or_none([r["_sold"] for r in rows]),
            "price_per_100":     _median_or_none(per_100s),
            "price_per_100_uom": uom_for_per100,
            "reviews":           _sum_or_none([r["_reviews"] for r in rows]),
            "avg_rating":        _mean_or_none([r["_rating"] for r in rows]),
            "top_products":      top_products,
            "platforms":         sorted({r.get("platform") for r in rows if r.get("platform")}),
        })

    out.sort(key=lambda r: (r["sales_volume"] is None, -(r["sales_volume"] or 0)))
    return out


# ═════════════════════════════════════════════════════════════════════════════
# LAYER C — workbook
# ═════════════════════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill("solid", fgColor="0F1E35")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BODY_ALIGN   = Alignment(horizontal="left", vertical="top",   wrap_text=True)
_BODY_NUM_ALIGN = Alignment(horizontal="right", vertical="top")


def _fmt_volume(val: float | None, uom: str | None) -> str:
    if val is None or uom is None:
        return "—"
    if uom == "ml" and val >= 1000:
        return f"{val/1000:.2f} L"
    if uom == "g"  and val >= 1000:
        return f"{val/1000:.2f} kg"
    return f"{val:.0f} {uom}"


def _fmt_per_100(val: float | None, uom: str | None) -> str:
    if val is None or uom is None:
        return "—"
    return f"Rp {int(round(val)):,} / 100{uom}".replace(",", ".")


def _fmt_top_products(items: list[dict]) -> str:
    if not items:
        return "—"
    lines = []
    for i, it in enumerate(items, start=1):
        sold_str = f"{int(it.get('sold') or 0):,}".replace(",", ".")
        title    = (it.get("title") or "").strip()
        if len(title) > 80:
            title = title[:78] + "…"
        lines.append(f"{i}. {title} — {sold_str} sold")
    return "\n".join(lines)


def _fmt_idr(val: float | None) -> str:
    if val is None:
        return "—"
    return f"Rp {int(round(val)):,}".replace(",", ".")


def _fmt_rating(val: float | None) -> str:
    if val is None:
        return "—"
    return f"{val:.2f}"


def _fmt_int(val: int | None) -> str:
    if val is None:
        return "—"
    return f"{int(val):,}".replace(",", ".")


def _autosize(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = max(14, min(48, len(h) + 4))


def _write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill  = _HEADER_FILL
        c.font  = _HEADER_FONT
        c.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _products_sheet(wb: Workbook, product_rows: list[dict]) -> None:
    """One row per tracked product (brand × flavour). Sorted by Sales Volume."""
    ws = wb.create_sheet("Products")
    headers = [
        "Product", "Sales Volume", "Unit Price per 100ml/g",
        "Top Products (top 3 by sold)", "Reviews", "# Listings", "Platforms",
    ]
    _write_header(ws, headers)
    for r in product_rows:
        ws.append([
            r["product_label"],
            _fmt_int(r["sales_volume"]),
            _fmt_per_100(r["price_per_100"], r["price_per_100_uom"]),
            _fmt_top_products(r["top_products"]),
            _fmt_int(r["reviews"]),
            r["n_listings"],
            ", ".join(r["platforms"]) if r["platforms"] else "—",
        ])
    # Style: wrap "Top Products" column, give it more width and taller rows
    # so the 3-line top list is fully visible.
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 60
        for c in row:
            c.alignment = _BODY_NUM_ALIGN if isinstance(c.value, (int, float)) else _BODY_ALIGN
    _autosize(ws, headers)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["D"].width = 70   # Top Products needs room


def _raw_sheet(wb: Workbook, parsed: list[dict]) -> None:
    ws = wb.create_sheet("Raw Listings")
    headers = [
        "Platform", "Brand", "Title", "Shop", "Official",
        "Flavour (parsed)", "Container (parsed)", "Total Units (parsed)",
        "Unit Volume (parsed)", "Total Volume (parsed)",
        "Listing Price (IDR)", "Per-Unit Price (IDR)",
        "Rating", "Reviews", "Sold", "URL",
    ]
    _write_header(ws, headers)
    for r in parsed:
        ws.append([
            r.get("platform"),
            r.get("brand_name"),
            r.get("title"),
            r.get("shop_name"),
            "Yes" if r.get("is_official_store") else "No",
            r.get("_flavour") or "—",
            r.get("_container"),
            r.get("_total_units"),
            _fmt_volume(r.get("_unit_volume"), r.get("_unit_volume_uom")),
            _fmt_volume(r.get("_total_volume"), r.get("_unit_volume_uom")),
            _fmt_idr(r.get("_listing_price")),
            _fmt_idr(r.get("_per_unit_price")),
            _fmt_rating(r.get("_rating")),
            _fmt_int(r.get("_reviews")),
            _fmt_int(r.get("_sold")),
            r.get("url"),
        ])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = _BODY_NUM_ALIGN if isinstance(c.value, (int, float)) else _BODY_ALIGN
    _autosize(ws, headers)


def _notes_sheet(wb: Workbook, n_listings: int, brand_filter: str | None) -> None:
    ws = wb.create_sheet("Notes")
    lines = [
        ("Competitor Analysis Export — How to read this file", True),
        ("", False),
        (f"Source: ecom_listings table, {n_listings} listing(s) "
         + (f"filtered to brand '{brand_filter}'." if brand_filter else "across all brands in this project."),
         False),
        ("", False),
        ("Sheets", True),
        ("• Products     — one row per tracked product (brand × flavour), sorted by Sales Volume desc. The user-specified brand + flavour from the scrape config win; regex parsing is only a fallback for legacy listings.", False),
        ("• Raw Listings — every individual scraped listing with the parser's output, for spot-checking.", False),
        ("", False),
        ("Columns", True),
        ("• Sales Volume          — sum of sold_count (from the actor's historicalSoldEstimated). Used as the sort key — most popular first.", False),
        ("• Unit Price per 100ml/g — median of (listing_price / total_units / unit_volume * 100) across the listings. Liquids report in ml, solids in g (majority UOM wins per group; mixed-UOM rows are excluded). '—' when no listing in the group has a parseable volume.", False),
        ("• Top Products          — top 3 listings within this product, ranked by sold count. Format: '1. <title> — N sold'. Quickest way to see which specific SKU is driving the brand's volume.", False),
        ("• Reviews               — sum of reviewCount across listings. Some actors don't expose a reviews field — falls back to '—'.", False),
        ("• # Listings            — count of validated listings (title-validated against brand + flavour at scrape time).", False),
        ("", False),
        ("Parser caveats", True),
        ("• Brand + flavour validation happens at scrape time — only listings whose TITLE contains ALL brand tokens AND ALL flavour tokens (case-insensitive) are kept. This filters out off-brand bleed from Shopee's loose-relevance search.", False),
        ("• Bundle / pack-count regex: 'isi N', 'N pcs/pack/sachet', 'Nx', 'xN', '1 lusin' (=12), 'N lusin' (=N×12), 'renceng N'. No signal → total_units defaults to 1.", False),
        ("• Volume regex: Nml / N L / Ng / N kg. L→×1000 (ml), kg→×1000 (g). Liquid vs solid never mixed in one aggregate.", False),
        ("• Parsed fields run FRESH each export. Phase 2 will persist them back into ecom_listings so repeated exports are faster and reproducible.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12)
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 110


def build_ecom_workbook(
    rows: list[dict],
    brand_filter: str | None = None,
    shop_filter: str | None = "all",
    specific_shops: list[str] | None = None,
) -> io.BytesIO:
    """Top-level entry: filter by shop → parse → aggregate-per-product → emit xlsx.
    shop_filter ∈ {all, official_only, non_official_only, specific_shops}."""
    rows          = _apply_export_shop_filter(rows, shop_filter, specific_shops)
    parsed        = [parse_listing(r) for r in rows]
    product_rows  = aggregate_by_product(parsed)

    wb = Workbook()
    wb.remove(wb.active)   # drop the default empty sheet
    _products_sheet(wb, product_rows)
    _raw_sheet(wb,      parsed)
    _notes_sheet(wb, len(rows), brand_filter)

    buf = io.BytesIO()
    wb.save(buf)
    return buf
