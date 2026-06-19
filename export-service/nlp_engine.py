"""
nlp_engine.py — NLP analysis engine + Excel compiler.

BASE_* constants are sourced from the original Cimory_Insta_CommentAnalysis_1_0.py.
They are the permanent floor for all projects — project config in Supabase ADDS to
these, never replaces them.

All BASE_* constants are exported so appv2.py can import them for pre-filling the
NLP Settings UI text areas with the full dictionary on first load.
"""
import io
import re
import difflib
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# 1. BASE DICTIONARIES  (complete — sourced from Cimory_Insta_CommentAnalysis_1_0.py)
# =============================================================================

BASE_PRODUCT = (
    "cimory, cimori, yogurt, susu, milk, squeeze, pouch, matcha, marie, biscuit, "
    "chocolate, strawberry, blueberry, mangga, mango, peach, taro, original, plain, "
    "gula, less sugar, tanpa gula, lactose, bebas lactose, lactose free, rasa, botol, "
    "drink, taste, flavour, flavor, sweetness, bottle, stroberi, choco mint, "
    "chocolate mint, coklat mint, sea salt, tanpa gula tambahan, stevia, biskuit, "
    "coklat, almond, laktosa, bebas laktosa, hazelnut, tiramisu, banyak rasa, "
    "creamy, light, uht, cashew"
)

BASE_INFLUENCER = (
    "cantik, ganteng, outfit, baju, spill, kak, kakak, abang, teh, teteh, mba, "
    "mbak, mas, rambut, makeup, lucu, gemes, gemoy, keren, idol"
)

BASE_POSITIVE = (
    "enak, mantap, suka, nagih, borong, segar, seger, halal, asli, murah, bagus, "
    "cocok, sip, boleh, rekomendasi, rekomen, love, good, best, fresh, tasty, "
    "delicious, yummy, worth, amazing, awesome, nice, perfect, favorite, favourite, "
    "recommend, like, happy, happier, great, excellent, fantastic, wonderful, obsessed, "
    "cool, enjoy, enjoyed, superb, brilliant, glad, addictive, kesukaan, mewah, nikmat, "
    "kriuk, kumplit, nostalgia, booster, cinta, juara"
)

BASE_NEGATIVE = (
    "mahal, kecewa, asem, kecut, aneh, bosen, kurang, cair, manis, eneg, pusing, "
    "zonk, boong, jelek, basi, bad, expensive, pricey, overpriced, awful, terrible, "
    "disgusting, sour, hate, boring, worst, sad, mad, angry, upset"
)

BASE_HEALTH = (
    "protein, tinggi protein, sehat, healthy, mindful, diet, nutrisi, gizi, "
    "kalori, rendah, low sugar"
)

BASE_INTENT = (
    "beli, dimana, indomaret, alfamart, stock, stok, pesen, order, harga, cari, "
    "ready, toko, supermarket, checkout, keranjang, mau, coba, cobain, nyoba, "
    "nyobain, penasaran, kepo, pengen, tertarik"
)

BASE_NEGATION_INTENSIFIERS = (
    "tidak, bukan, jangan, belum, "
    "banget, paling, beneran, really, very, super, parah, sekali, amat, sangat, "
    "juara, harus, hrs"
)

BASE_SPAM = (
    "demi allah, like komen, orang tercepat, fb serius, follback, cek dm, "
    "p adu jam, titip sendal, hadir, absen, subs"
)

BASE_REACTION = "tertawa, lol, wow, thinking"

BASE_EMOJI_POSITIVE = "❤️, 💖, 🔥, 👍, 😋, 😍, 👏, 🤤, 🥛, 💯, ✨"
BASE_EMOJI_NEGATIVE = "🤮, 💩, 😡, 👎, 😒, 🤡, 💀, 🙃, 🤢, 🗿"

# PR Crisis defaults (Cimory Fresh Milk context)
BASE_PR_NAME     = "Fresh Milk"
BASE_PR_PHRASES  = "33%, 33 persen, cuma 33, 33 doang, fresh milk, susu segar"
BASE_PR_KEYWORDS = "kandungan, komposisi, campuran, kadar, persentase"
BASE_PR_TRIGGERS = "dikit, sedikit, cuma, doang, kurang, air"

BASE_THEME_MAP = {
    "Lactose Focus":        "lactose, laktosa, bebas laktosa, lactose free",
    "Sugar / Sweetness":    "gula, manis, less sugar, stevia, tanpa gula, kemanisan",
    "Packaging / Form":     "botol, bottle, pouch, squeeze",
    "Price Sensitivity":    "harga, mahal, murah, price, murmer, promo, diskon",
    "Diet & Health":        "diet, protein, kalori, sehat, nutrisi",
}

# Slang + typo corrections merged into one map
BASE_SLANG_MAP = {
    # Intensifiers / Fillers
    "yg": "yang",    "bgt": "banget",  "bngt": "banget",   "bngtt": "banget",
    "benget": "banget", "bangat": "banget", "bngit": "banget",
    "blm": "belum",  "jg": "juga",     "udh": "sudah",     "udah": "sudah",
    "krn": "karena", "karna": "karena","tp": "tetapi",     "tapi": "tetapi",
    # Negations
    "gk": "tidak",   "ga": "tidak",    "gak": "tidak",     "engga": "tidak",
    "ngk": "tidak",  "nggk": "tidak",  "tdk": "tidak",
    # Common shortenings
    "gpp": "tidak apa-apa", "dr": "dari", "dlm": "dalam", "dpt": "dapat",
    "klo": "kalau",  "kl": "kalau",    "lu": "kamu",
    "gw": "saya",    "aku": "saya",    "sy": "saya",
    # Commerce
    "reco": "rekomendasi", "recomen": "rekomendasi", "rekomen": "rekomendasi",
    "brg": "barang",  "mhl": "mahal",  "murmer": "murah meriah",
    "brp": "berapa",  "brpa": "berapa","hargax": "harganya", "hrg": "harga",
    "nyari": "cari",  "bli": "beli",   "pesen": "pesan",   "ongkir": "ongkos kirim",
    # Questions / Expressions
    "knp": "kenapa", "gmn": "bagaimana","gimana": "bagaimana",
    "kyk": "seperti","kayak": "seperti",
    "bgs": "bagus",  "bgus": "bagus",  "bagoss": "bagus",
    "jgn": "jangan", "bener": "benar", "bnr": "benar",
    "gtu": "begitu", "gitu": "begitu",
    "skrg": "sekarang","trs": "terus", "sm": "sama",   "bs": "bisa",
    "bkn": "bukan",  "emg": "memang",  "sih": "sih",   "dong": "dong", "donk": "dong",
    "maf": "maaf",   "sat": "saat",    "jd": "jadi",   "pgn": "pengen",
    "tmbh": "tambah","btl": "botol",   "kdg": "kadang","hmmm": "thinking",
    # Laughter
    "wkwk": "tertawa","wkwkwk": "tertawa","xixixi": "tertawa",
    "haha": "tertawa","hahaha": "tertawa","hehe": "tertawa","hihi": "tertawa",
    # Brand / product typos
    "cimori": "cimory",   "chimory": "cimory",   "cymory": "cimory",
    "yoghurt": "yogurt",  "yogurtnya": "yogurt",
    "mantaap": "mantap",  "manteb": "mantap",    "mantab": "mantap",
    "strowberi": "strawberry","stroberi": "strawberry","strawberi": "strawberry",
    "coklat": "chocolate","cokelat": "chocolate",
}

# Module-level constants — not per-project configurable
NEGATION_WORDS = {"tidak", "bukan", "jangan", "belum"}
INTENSIFIERS   = {"banget","paling","beneran","really","very","super","parah",
                  "sekali","amat","sangat","juara","harus","hrs"}
INTENT_WORDS   = {"beli","dimana","indomaret","alfamart","stock","stok","pesen",
                  "order","harga","cari","ready","toko","supermarket","checkout",
                  "keranjang","mau","coba","cobain","nyoba","nyobain","penasaran",
                  "kepo","pengen","tertarik"}
SPAM_PHRASES   = {"demi allah","like komen","orang tercepat","fb serius","follback",
                  "cek dm","p adu jam","titip sendal","hadir","absen","subs"}
REACTION_WORDS = {"tertawa","lol","wow","thinking"}
EMOJI_POSITIVE = ["❤️","💖","🔥","👍","😋","😍","👏","🤤","🥛","💯","✨"]
EMOJI_NEGATIVE = ["🤮","💩","😡","👎","😒","🤡","💀","🙃","🤢","🗿"]


# =============================================================================
# 2. CONFIG MERGE HELPERS
# =============================================================================
def _csv_to_set(text):
    return {w.strip().lower() for w in str(text or "").split(",") if w.strip()}

def _merge_kw(config_val, base_str):
    """BASE ∪ project config — config adds to base, never replaces it."""
    return _csv_to_set(base_str) | _csv_to_set(config_val)

def _merge_slang(config_val):
    merged = dict(BASE_SLANG_MAP)
    if isinstance(config_val, list):
        for s in config_val:
            if isinstance(s, dict) and "Original" in s and "Replacement" in s:
                k = str(s["Original"]).strip().lower()
                v = str(s["Replacement"]).strip().lower()
                if k and v:
                    merged[k] = v
    elif isinstance(config_val, dict):
        for k, v in config_val.items():
            merged[str(k).strip().lower()] = str(v).strip().lower()
    return merged

def _merge_themes(config_val):
    merged_src = dict(BASE_THEME_MAP)
    if isinstance(config_val, list):
        for t in config_val:
            if isinstance(t, dict) and "Theme" in t:
                name = t["Theme"]
                kws  = t.get("Keywords", "")
                merged_src[name] = (merged_src[name] + ", " + kws) if name in merged_src else kws
    elif isinstance(config_val, dict):
        for name, kws in config_val.items():
            merged_src[name] = (merged_src[name] + ", " + kws) if name in merged_src else kws
    return {name: _csv_to_set(kws) for name, kws in merged_src.items()}

def _build_kw_context(config):
    """Build keyword context ONCE per Excel generation, not per comment."""
    slang  = _merge_slang(config.get("slang_map"))
    themes = _merge_themes(config.get("theme_map"))

    all_valid = (
        set(slang.keys()) | set(slang.values()) |
        _merge_kw(config.get("product_keywords"),    BASE_PRODUCT) |
        _merge_kw(config.get("influencer_keywords"), BASE_INFLUENCER) |
        _merge_kw(config.get("positive_words"),      BASE_POSITIVE) |
        _merge_kw(config.get("negative_words"),      BASE_NEGATIVE) |
        _merge_kw(config.get("health_keywords"),     BASE_HEALTH) |
        INTENT_WORDS | NEGATION_WORDS | INTENSIFIERS
    )
    for kw_set in themes.values():
        all_valid |= kw_set

    brand_raw = config.get("brand_handle", "cimory") or "cimory"

    # ── PR Crisis alerts — supports multiple named alerts ──────────────────
    # Stored as config["pr_alerts"]: a list of {name, phrases, keywords, triggers}.
    # Falls back to the legacy flat pr_name/pr_phrases/pr_keywords/pr_triggers
    # columns if pr_alerts hasn't been saved yet, so existing projects keep
    # working without needing to re-save their config.
    pr_alerts_raw = config.get("pr_alerts")
    if not pr_alerts_raw:
        legacy_present = any(config.get(f) for f in ("pr_name", "pr_phrases", "pr_keywords", "pr_triggers"))
        pr_alerts_raw = [{
            "name":     config.get("pr_name", BASE_PR_NAME),
            "phrases":  config.get("pr_phrases", BASE_PR_PHRASES),
            "keywords": config.get("pr_keywords", BASE_PR_KEYWORDS),
            "triggers": config.get("pr_triggers", BASE_PR_TRIGGERS),
        }] if legacy_present else [{
            "name": BASE_PR_NAME, "phrases": BASE_PR_PHRASES,
            "keywords": BASE_PR_KEYWORDS, "triggers": BASE_PR_TRIGGERS,
        }]
    if isinstance(pr_alerts_raw, dict):
        pr_alerts_raw = [pr_alerts_raw]

    pr_alerts = []
    for a in pr_alerts_raw:
        name = str(a.get("name", a.get("Alert Name", "PR Alert"))).strip() or "PR Alert"
        phrases = [w.strip().lower() for w in
                   str(a.get("phrases", a.get("Exact Phrases", ""))).split(",") if w.strip()]
        pr_alerts.append({
            "name":     name,
            "phrases":  phrases,
            "kws":      _csv_to_set(a.get("keywords", a.get("Keywords", ""))),
            "negs":     _csv_to_set(a.get("triggers", a.get("Triggers", ""))),
        })

    return {
        "product":    _merge_kw(config.get("product_keywords"),    BASE_PRODUCT),
        "influencer": _merge_kw(config.get("influencer_keywords"), BASE_INFLUENCER),
        "positive":   _merge_kw(config.get("positive_words"),      BASE_POSITIVE),
        "negative":   _merge_kw(config.get("negative_words"),      BASE_NEGATIVE),
        "health":     _merge_kw(config.get("health_keywords"),     BASE_HEALTH),
        "slang":      slang,
        "themes":     themes,
        "all_valid":  list(all_valid),
        "brand":      "@" + brand_raw.lower().lstrip("@"),
        "pr_alerts":  pr_alerts,
        # Legacy single-name key kept for any code path that still reads kw["pr_name"]
        # directly (e.g. an old cached Excel column header) — uses the first alert.
        "pr_name":    pr_alerts[0]["name"] if pr_alerts else BASE_PR_NAME,
    }


# =============================================================================
# 3. PER-COMMENT ANALYSIS
# =============================================================================
def analyze_comment(text, commenter, kol_username, kw):
    raw_lower = str(text).lower()
    commenter = str(commenter).strip().lower()
    kol       = str(kol_username).strip().lower()

    # Early exits
    if commenter == kol:
        return {"sentiment": "Creator/Ignored", "subject": "Creator Reply",
                "themes": "", "intent": 0, "pr": "No",
                "pos_words": 0, "health": 0, "brand": 0,
                "unk": [], "pos_emo": "", "neg_emo": ""}
    if any(spam in raw_lower for spam in SPAM_PHRASES):
        return {"sentiment": "Spam/Ignored", "subject": "Spam/Bait",
                "themes": "", "intent": 0, "pr": "No",
                "pos_words": 0, "health": 0, "brand": 0,
                "unk": [], "pos_emo": "", "neg_emo": ""}

    # PR check — evaluate every configured alert, collect all that fire
    raw_words = raw_lower.split()
    fired_alerts = []
    for alert in kw["pr_alerts"]:
        phrase_hit = any(p in raw_lower for p in alert["phrases"]) if alert["phrases"] else False
        kws, negs = alert["kws"], alert["negs"]
        if kws and negs:
            combo_hit = any(w in kws for w in raw_words) and any(w in negs for w in raw_words)
        elif kws:
            combo_hit = any(w in kws for w in raw_words)
        elif negs:
            combo_hit = any(w in negs for w in raw_words)
        else:
            combo_hit = False
        if phrase_hit or combo_hit:
            fired_alerts.append(alert["name"])
    is_pr   = len(fired_alerts) > 0
    pr_flag = "YES" if is_pr else "No"

    # Emoji detection
    pos_e_list = [e for e in EMOJI_POSITIVE if e in text]
    neg_e_list = [e for e in EMOJI_NEGATIVE if e in text]
    pos_e, neg_e = len(pos_e_list), len(neg_e_list)

    # Word normalisation
    words        = raw_lower.split()
    slang_map    = kw["slang"]
    all_valid    = kw["all_valid"]
    cleaned, unk = [], []

    for w in words:
        if w in slang_map:
            cleaned.append(slang_map[w])
        elif w in all_valid:
            cleaned.append(w)
        else:
            shrunk = re.sub(r"([a-z])\1+", r"\1", w)
            if shrunk in slang_map:
                cleaned.append(slang_map[shrunk])
            elif shrunk in all_valid:
                cleaned.append(shrunk)
            elif len(w) >= 5:
                matches = difflib.get_close_matches(w, all_valid, n=1, cutoff=0.82)
                cleaned.append(matches[0] if matches else w)
                if not matches and len(w) > 3 and not w.startswith("@"):
                    unk.append(w)
            else:
                cleaned.append(w)
                if len(w) > 3 and not w.startswith("@"):
                    unk.append(w)

    # Scoring
    pos_score = neg_score = 0
    intent_flag = 1 if any(w in INTENT_WORDS   for w in cleaned) else 0
    health_flag = 1 if any(w in kw["health"]   for w in cleaned) else 0
    brand_flag  = 1 if kw["brand"] in raw_lower else 0

    for i, w in enumerate(cleaned):
        mult   = 1.5 if (i > 0 and cleaned[i-1] in INTENSIFIERS) else 1
        is_neg = ((i > 0 and cleaned[i-1] in NEGATION_WORDS) or
                  (i > 1 and cleaned[i-2] in NEGATION_WORDS))
        if w in kw["positive"]:
            if is_neg: neg_score += mult
            else:      pos_score += mult
        elif w in kw["negative"]:
            if is_neg: pos_score += mult
            else:      neg_score += mult

    net = pos_score - neg_score
    if net > 0 and neg_e > 0:   net = -1
    elif net < 0 and pos_e > 2: net = 0
    net += (pos_e * 0.5) - (neg_e * 0.5)

    sent = "Positive" if net > 0 else ("Negative" if net < 0 else "Neutral")
    if is_pr: sent = "Negative (Forced)"

    # Themes
    themes = list(fired_alerts) if is_pr else []
    for t_name, t_kws in kw["themes"].items():
        if any(w in t_kws for w in cleaned):
            themes.append(t_name)

    # Subject classification
    m_prod = any(w in kw["product"]    for w in cleaned)
    m_inf  = any(w in kw["influencer"] for w in cleaned)
    if   m_prod and not m_inf:  subj = "Product Focused"
    elif m_inf  and not m_prod: subj = "Creator Focused"; intent_flag = 0
    elif m_prod and m_inf:      subj = "Mixed Mention"
    elif pos_e > 0 and not cleaned: subj = "General/Emoji Affinity"
    else:                           subj = "Unclear/General"

    return {
        "sentiment": sent,   "subject":  subj,
        "themes":    ", ".join(themes),
        "intent":    intent_flag, "pr":    pr_flag,
        "pos_words": pos_score,   "health": health_flag, "brand": brand_flag,
        "unk":       unk,   "pos_emo": "".join(pos_e_list), "neg_emo": "".join(neg_e_list),
    }


# =============================================================================
# 4. EXCEL COMPILER
# =============================================================================
def _hdr(ws):
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def _widen(ws, spec):
    for col, w in spec:
        ws.column_dimensions[col].width = w

def generate_nlp_excel(df_raw, config):
    df = df_raw.rename(columns={
        "comment_text":        "text",
        "commenter_username":  "ownerUsername",
        "influencer_username": "KOL Username",
        "video_url":           "postUrl",
    })

    kw = _build_kw_context(config)

    results = [
        analyze_comment(row.get("text", ""), row.get("ownerUsername", ""),
                        row.get("KOL Username", ""), kw)
        for _, row in df.iterrows()
    ]

    df["Calculated_Sentiment"] = [r["sentiment"]    for r in results]
    df["Conversation_Subject"] = [r["subject"]      for r in results]
    df["Detected_Themes"]      = [r["themes"]       for r in results]
    df["Purchase_Intent"]      = [r["intent"]       for r in results]
    df["PR_Crisis_Flag"]       = [r["pr"]           for r in results]
    df["pos_words"]            = [r["pos_words"]    for r in results]
    df["health"]               = [r["health"]       for r in results]
    df["brand"]                = [r["brand"]        for r in results]
    df["unk"]                  = [", ".join(r["unk"]) for r in results]
    df["pos_emo"]              = [r["pos_emo"]      for r in results]
    df["neg_emo"]              = [r["neg_emo"]      for r in results]

    noise    = ["Spam/Ignored", "Creator/Ignored", "Reaction/Ignored"]
    df_noise = df[df["Calculated_Sentiment"].isin(noise)]
    df_valid = df[~df["Calculated_Sentiment"].isin(noise)]
    pr_name  = kw["pr_name"]

    # Executive Summary
    vid_summary = []
    for post_url, group in df.groupby("postUrl"):
        total_raw    = len(group)
        creator_name = group["KOL Username"].iloc[0] if len(group) else "Unknown"
        v_group      = group[~group["Calculated_Sentiment"].isin(noise)]
        v_count      = len(v_group)
        if v_count == 0:
            continue

        pos_c     = sum(v_group["Calculated_Sentiment"] == "Positive")
        neg_c     = sum(v_group["Calculated_Sentiment"].str.contains("Negative"))
        prod_rel  = v_group["Conversation_Subject"].isin(["Product Focused","Mixed Mention"]).sum() / v_count * 100
        net_sent  = (pos_c - neg_c) / v_count * 100
        intent_rt = v_group["Purchase_Intent"].sum() / v_count * 100
        health_rt = v_group["health"].sum()           / v_count * 100
        brand_cnt = v_group["brand"].sum()
        avg_pos   = v_group["pos_words"].sum()        / v_count
        noise_rt  = (total_raw - v_count) / total_raw * 100 if total_raw else 0
        unique_c  = v_group["ownerUsername"].nunique()
        v_score   = (max(0, min(100, (net_sent + 100) / 2)) * 0.40 +
                     intent_rt * 0.40 + prod_rel * 0.20)
        integrity = (
            "SEVERE WARNING: ≤5 Valid Comments (NOT RELIABLE)" if v_count <= 5 else
            "Unreliable — Under 30 Valid Comments"             if v_count <= 30 else
            "Reliable Sample"
        )
        slang_raw    = ", ".join(s for s in v_group["unk"] if s)
        unique_slang = ", ".join(sorted(set(s.strip() for s in slang_raw.split(",") if s.strip())))

        vid_summary.append({
            "Creator Username":              creator_name,
            "Video URL":                     post_url,
            "Total Scraped":                 total_raw,
            "Valid Comments":                v_count,
            "Unique Commenters":             unique_c,
            "Sample Integrity":              integrity,
            "FINAL VIDEO SCORE (0-100)":     round(v_score, 1),
            f"🚨 {pr_name} Complaints":      int(v_group["PR_Crisis_Flag"].eq("YES").sum()),
            "Avg Positivity Density":        round(avg_pos, 2),
            "Brand Tag Count":               int(brand_cnt),
            "Health Focus Rate":             f"{round(health_rt, 1)}%",
            "Product Relevance Rate":        f"{round(prod_rel, 1)}%",
            "Net Sentiment Rate":            f"{round(net_sent, 1)}%",
            "Purchase Intent Rate":          f"{round(intent_rt, 1)}%",
            "Noise / Spam Rate":             f"{round(noise_rt, 1)}%",
            "New Slang Detected":            unique_slang,
        })

    df_exec = (
        pd.DataFrame(vid_summary).sort_values("FINAL VIDEO SCORE (0-100)", ascending=False)
        if vid_summary else pd.DataFrame(columns=["No Data"])
    )

    # Theme Trend
    themes_flat = ", ".join(t for t in df_valid["Detected_Themes"] if t)
    if themes_flat:
        unique_themes = set(themes_flat.split(", ")) - {""}
        theme_counts  = {
            t: sum(1 for ts in df_valid["Detected_Themes"] if ts and t in ts.split(", "))
            for t in unique_themes
        }
        df_themes = (pd.DataFrame(list(theme_counts.items()), columns=["Theme","Mentions"])
                     .sort_values("Mentions", ascending=False))
    else:
        df_themes = pd.DataFrame(columns=["Theme","Mentions"])

    # Creator Positivity
    creator_stats, total_valid_g, total_prod_pos_g = [], 0, 0
    for creator, cg in df_valid.groupby("KOL Username"):
        tot      = len(cg)
        prod_pos = len(cg[
            cg["Conversation_Subject"].isin(["Product Focused","Mixed Mention"]) &
            (cg["Calculated_Sentiment"] == "Positive")
        ])
        total_valid_g   += tot
        total_prod_pos_g += prod_pos
        creator_stats.append({
            "Creator":                      creator,
            "Total Valid Comments":          tot,
            "Product-Focused Positive":      prod_pos,
            "Positivity Conversion Rate":    f"{round(prod_pos/tot*100, 1)}%" if tot else "0%",
        })
    df_creator = (pd.DataFrame(creator_stats).sort_values("Total Valid Comments", ascending=False)
                  if creator_stats else pd.DataFrame(columns=["No Data"]))
    if not df_creator.empty:
        grand = pd.DataFrame([{
            "Creator":                   "GRAND TOTAL",
            "Total Valid Comments":       total_valid_g,
            "Product-Focused Positive":   total_prod_pos_g,
            "Positivity Conversion Rate": f"{round(total_prod_pos_g/max(total_valid_g,1)*100, 1)}%",
        }])
        df_creator = pd.concat([grand, df_creator], ignore_index=True)

    comment_cols = ["postUrl","ownerUsername","KOL Username","text",
                    "Calculated_Sentiment","Conversation_Subject","Detected_Themes",
                    "pos_emo","neg_emo"]
    sheets = {
        "Executive Summary":          df_exec,
        "Theme Trend Analysis":       df_themes,
        "Creator Product Positivity": df_creator,
        f"{pr_name} & PR Mentions":   df_valid[df_valid["PR_Crisis_Flag"] == "YES"][comment_cols],
        "Product-Specific Feedback":  df_valid[
            df_valid["Conversation_Subject"].isin(["Product Focused","Mixed Mention"])
        ][comment_cols],
        "Positive & Neutral":         df_valid[
            df_valid["Calculated_Sentiment"].isin(["Positive","Neutral"])
        ][comment_cols],
        "Negative Comments":          df_valid[
            df_valid["Calculated_Sentiment"].str.contains("Negative")
        ][comment_cols],
        "Filtered Noise":             df_noise[
            ["postUrl","ownerUsername","text","Calculated_Sentiment","pos_emo","neg_emo"]
        ],
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for s_name, s_data in sheets.items():
        if s_data.empty:
            s_data = pd.DataFrame(columns=["No Data"])
        ws = wb.create_sheet(title=s_name[:31])
        for r in dataframe_to_rows(s_data, index=False, header=True):
            ws.append(r)
        _hdr(ws)
        _widen(ws, [("A",30),("B",28),("C",20),("D",22),("E",20),("F",22),("G",26),("H",14),("I",14)])
        if "text" in s_data.columns:
            tc = openpyxl.utils.get_column_letter(s_data.columns.get_loc("text") + 1)
            ws.column_dimensions[tc].width = 70
        if "postUrl" in s_data.columns:
            uc = openpyxl.utils.get_column_letter(s_data.columns.get_loc("postUrl") + 1)
            ws.column_dimensions[uc].width = 55
        if s_name == "Executive Summary":
            ws.freeze_panes = "C2"
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column == 6 and "WARNING" in str(cell.value or ""):
                        cell.font = Font(color="FF0000", bold=True)
                    elif cell.column == 8 and isinstance(cell.value, (int,float)) and cell.value > 0:
                        cell.font = Font(bold=True, color="C00000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
