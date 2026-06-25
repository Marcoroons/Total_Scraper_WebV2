import re
"""
utils.py — shared helpers used across all tab modules and worker.
Covers: CSS injection, SMTP email, Excel generators, job-type config.
"""
import io
import os
import smtplib
from email.message import EmailMessage

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ─────────────────────────────────────────────────────────────────────────────
# JOB TYPE METADATA
# Drives column labels, metric selectors, duration estimates, and UI hints.
# ─────────────────────────────────────────────────────────────────────────────
JOB_META = {
    "Specific URLs (Video Stats)": {
        "input_cols":    ["KOL Username", "Video URL"],
        "rate_eligible": True,
        "metric_eligible": True,
        "duration":      "1–3 min per URL",
        "help": (
            "**What it does:** Pulls engagement metrics for specific posts you already know.\n\n"
            "**Input:** Paste the full post URL in *Video URL*. "
            "*KOL Username* is optional but helps with ER context."
        ),
    },
    "Profile Feed (Audit)": {
        "input_cols":    ["Profile Username or URL"],
        "rate_eligible": True,
        "metric_eligible": True,
        "duration":      "3–7 min per profile",
        "help": (
            "**What it does:** Scrapes recent posts from a creator's profile.\n\n"
            "**Input:** Enter a bare username (`johndoe`), `@johndoe`, "
            "or a full profile URL — any format works."
        ),
    },
    "Comments (Sentiment)": {
        "input_cols":    ["KOL Username", "Video URL"],
        "rate_eligible": False,
        "metric_eligible": False,
        "duration":      "4–10 min per video",
        "help": (
            "**What it does:** Scrapes comments and runs NLP sentiment scoring.\n\n"
            "**Input:** *Video URL* is required. *KOL Username* filters "
            "out creator replies so they don't skew sentiment."
        ),
    },
    "Trend Discovery (Hashtag)": {
        "input_cols":    ["Hashtag(s)"],
        "rate_eligible": False,
        "metric_eligible": False,
        "duration":      "5–15 min per hashtag",
        "help": (
            "**What it does:** Scrapes top public posts under a hashtag to reveal "
            "what content structure, length, and sounds are winning.\n\n"
            "**Input:** Enter one or more hashtags in *Hashtag(s)*, e.g. `OOTD` or "
            "`#OOTD, #Fashion`. No `#` needed — it's stripped automatically. "
            "Separate multiple hashtags with commas."
        ),
    },
    "Trend Discovery (User Profile)": {
        "input_cols":    ["Competitor Profile"],
        "rate_eligible": False,
        "metric_eligible": False,
        "duration":      "3–7 min per profile",
        "help": (
            "**What it does:** Analyses a competitor's or reference creator's top-performing "
            "videos to extract structural patterns.\n\n"
            "**Input:** Enter the username or full profile URL in *Competitor Profile*."
        ),
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT CSS  (inject once from appv2.py)
# ─────────────────────────────────────────────────────────────────────────────
APP_CSS = """
<style>
/* Metric cards */
[data-testid="stMetric"] {
    background: rgba(31,78,120,0.12);
    border: 1px solid rgba(31,78,120,0.35);
    border-radius: 10px;
    padding: 14px 18px;
}
/* Primary button */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg,#1F4E78 0%,#2E86AB 100%);
    border: none; border-radius: 6px; font-weight:600; color:#fff;
}
/* Tabs */
.stTabs [data-baseweb="tab"] { font-weight:600; font-size:13px; }
/* Info/warning boxes */
.stAlert { border-radius: 8px; }
/* Expander */
.streamlit-expanderHeader { font-weight:600; }
/* Dataframe */
[data-testid="stDataFrame"] { border-radius:8px; }
/* Status pill helpers (use via st.markdown) */
.pill-green { background:#d4edda; color:#155724; padding:2px 10px; border-radius:12px; font-size:12px; font-weight:600; }
.pill-yellow{ background:#fff3cd; color:#856404; padding:2px 10px; border-radius:12px; font-size:12px; font-weight:600; }
.pill-red   { background:#f8d7da; color:#721c24; padding:2px 10px; border-radius:12px; font-size:12px; font-weight:600; }
</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────
def _get_bot_creds():
    email    = os.environ.get("BOT_EMAIL")
    password = os.environ.get("BOT_APP_PASSWORD")
    if not email or not password:
        try:
            email    = st.secrets.get("BOT_EMAIL")
            password = st.secrets.get("BOT_APP_PASSWORD")
        except Exception:
            pass
    return email, password


def dispatch_email(msg: EmailMessage) -> tuple[bool, str]:
    """
    Send a pre-built EmailMessage via Gmail SMTP_SSL.
    Returns (success: bool, message: str).
    Uses SMTP_SSL port 465 — more reliable than STARTTLS in containerised envs.
    """
    bot_email, bot_password = _get_bot_creds()
    if not bot_email or not bot_password:
        return False, "BOT_EMAIL or BOT_APP_PASSWORD not set in environment."
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(bot_email, bot_password)
            smtp.send_message(msg)
        return True, "Email sent successfully."
    except Exception as e:
        return False, f"SMTP error: {e}"


def send_report_email(
    from_email: str,
    to_csv: str,
    subject: str,
    body: str,
    excel_buffer,
    filename: str,
) -> tuple[bool, str]:
    """Convenience wrapper — builds + dispatches a report email."""
    bot_email, _ = _get_bot_creds()
    if not bot_email:
        return False, "BOT_EMAIL not configured."

    msg             = EmailMessage()
    msg["Subject"]  = subject
    msg["From"]     = f"Total Scraper System <{bot_email}>"
    msg["To"]       = ", ".join(e.strip() for e in to_csv.split(",") if e.strip())
    msg["Reply-To"] = from_email
    msg.set_content(body)

    excel_buffer.seek(0)
    msg.add_attachment(
        excel_buffer.read(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
    return dispatch_email(msg)


def send_invite_email(inviter_email: str, recipient_email: str, team_name: str) -> tuple[bool, str]:
    bot_email, _ = _get_bot_creds()
    if not bot_email:
        return False, "BOT_EMAIL not configured."

    msg             = EmailMessage()
    msg["Subject"]  = f"🤝 You've been invited to join '{team_name}' on Total Scraper"
    msg["From"]     = f"Total Scraper System <{bot_email}>"
    msg["To"]       = recipient_email
    msg["Reply-To"] = inviter_email
    msg.set_content(
        f"Hello!\n\n{inviter_email} has invited you to collaborate in the "
        f"'{team_name}' workspace on Total Scraper Web.\n\n"
        f"To join:\n1. Go to the app.\n2. Create an account using {recipient_email}.\n"
        f"3. Use the Secret Invite Code provided by your admin.\n\nWelcome!\nTotal Scraper Bot"
    )
    return dispatch_email(msg)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _fmt(val):
    """Return 'Hidden' for -1 metrics, empty for NaN."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return "Hidden" if val == -1 else val


def _apply_header(ws):
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _widen(ws, spec: list[tuple[str, int]]):
    for col, w in spec:
        ws.column_dimensions[col].width = w


def generate_video_stats_excel(df_raw: pd.DataFrame, is_tiktok: bool = False,
                               calc_metrics=None, raw_metrics=None) -> io.BytesIO:
    from openpyxl.utils import get_column_letter
    df = df_raw.copy()
    df["Likes"]    = df["likes"].apply(_fmt)
    df["Comments"] = df["comments"].apply(_fmt)
    df["Shares"]   = df["shares"].apply(_fmt)
    df["Plays"]    = df["play_count"].apply(_fmt)

    # Calculated metrics chosen at export time. Only those derivable from
    # likes/comments/shares/play_count. VTR is intentionally dropped — there is no
    # separate view count (it was always "100%"); CPV needs a per-video rate the
    # video table doesn't carry.
    _CALC = {
        "Engagement Rate":    lambda r: (r["likes"] + r["comments"] + r["shares"]) / max(r["play_count"], 1) * 100,
        "Applause Rate":      lambda r: r["likes"] / max(r["play_count"], 1) * 100,
        "Virality Rate":      lambda r: r["shares"] / max(r["play_count"], 1) * 100,
        "Comment/View Ratio": lambda r: r["comments"] / max(r["play_count"], 1) * 100,
    }
    sel = [m for m in (calc_metrics or []) if m in _CALC]
    if not sel:
        sel = ["Engagement Rate"]

    # A post with no view count (a photo, or a privacy-restricted video) can't carry
    # a view-based rate — show "N/A" instead of a misleading divide-by-1 figure.
    def _metric(row, m):
        if "Hidden" in (row["Likes"], row["Comments"], row["Shares"], row["Plays"]):
            return "Hidden"
        if int(row.get("play_count") or 0) <= 0:
            return "N/A"
        return f"{_CALC[m](row):.2f}%"
    for m in sel:
        df[m] = df.apply(lambda r, _m=m: _metric(r, _m), axis=1)

    # Type column: reels/videos carry views, photos don't. Prefer a stored
    # content_type when present; otherwise infer from the presence of a play count.
    _IMG_CT = {"image", "graphimage", "photo", "sidecar", "graphsidecar", "carousel"}
    _VID_CT = {"video", "graphvideo", "reel", "clips", "igtv"}
    def _vtype(row):
        ct = str(row.get("content_type", "") or "").strip().lower()
        if ct in _IMG_CT:
            return "Image"
        if ct in _VID_CT:
            return "Video"
        return "Video" if int(row.get("play_count") or 0) > 0 else "Image"
    df["Type"] = df.apply(_vtype, axis=1)

    _sel_raw = set(raw_metrics) if raw_metrics else {"Likes", "Comments", "Shares"}
    raw_cols = [rc for rc in ("Likes", "Comments", "Shares") if rc in _sel_raw]
    src_cols = ["username", "video_url", "Type", "Plays"] + raw_cols + sel
    out = df[src_cols].copy()
    out.columns = ["Username", "Video URL", "Type", "Play Count"] + raw_cols + sel

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Video Stats")
        ws = writer.sheets["Video Stats"]
        _apply_header(ws)
        widths = [
            (get_column_letter(i),
             25 if c == "Username" else 55 if c == "Video URL" else 16 if c == "Play Count"
             else 10 if c == "Type" else 15)
            for i, c in enumerate(out.columns, 1)
        ]
        _widen(ws, widths)
        ws.freeze_panes = "C2"
    buf.seek(0)
    return buf


def generate_profile_audit_excel(
    df_raw: pd.DataFrame,
    is_tiktok: bool = False,
    sort_by: str = "Most Views",
    incl_top5: bool = True,
    incl_bot5: bool = False,
    limit: int = 0,
    calc_metrics=None,
    raw_metrics=None,
    rates=None,
    date_from: str = "",
    date_to: str = "",
    requested_usernames=None,
    layout: dict = None,
) -> io.BytesIO:
    """
    Compact Profile Feed (Audit) export.

    Layout (one row per creator):
      Fixed: KOL/Creator | Platform | # Videos | Avg Views | Most Views | Least Views
             Date of Most Viewed | Date of Least Viewed
      Optional: Top 5 Avg Views | Bottom 5 Avg Views
      Video columns: V1 | V2 | V3 | V4 | ...
        — Each cell value = plain view count (integer, easy to select/copy as a range)
        — Each cell carries an openpyxl Comment: "🔗 URL\n📅 Posted: YYYY-MM-DD"

    Sheet 2 "Full Details" — one row per video, readable table with URL + date columns.
    Sheet 3 "Export Notes" — platform disclaimers.
    """
    import re
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    df = df_raw.copy()
    platform_label = "TikTok" if is_tiktok else "Instagram"

    # Ensure required columns exist
    if "play_count" not in df.columns:
        df["play_count"] = 0
    if "post_url"   not in df.columns:
        df["post_url"]   = ""
    if "post_date"  not in df.columns:
        df["post_date"]  = ""
    if "username"   not in df.columns:
        df["username"]   = "unknown"

    df["play_count"] = pd.to_numeric(df["play_count"], errors="coerce").fillna(0).astype(int)
    df["post_url"]   = df["post_url"].fillna("").astype(str)
    df["post_date"]  = df["post_date"].fillna("").astype(str).str[:10]

    # ── Respect the chosen scrape window ────────────────────────────────────────
    # influencer_profiles accumulates EVERY scrape of a creator, so without this an
    # export would show out-of-window posts (e.g. June posts when you asked for
    # "up to 30 May") — making the date filter look ignored. Keep only dated posts
    # inside [date_from, date_to]; undated rows are dropped when a window is set
    # since they can't be confirmed in range.
    if date_from or date_to:
        _parsed = pd.to_datetime(df["post_date"], errors="coerce")
        _mask = _parsed.notna()
        if date_from:
            _mask &= _parsed >= pd.to_datetime(date_from, errors="coerce")
        if date_to:
            _mask &= _parsed <= pd.to_datetime(date_to, errors="coerce")
        df = df[_mask].reset_index(drop=True)

    # ── Classify each post: Video (reel — has a view count) vs Image (photo/
    # carousel — no view count). View-based metrics only make sense for videos, so
    # the summary ranks videos only and image posts list view metrics as N/A.
    # content_type (captured by the worker) is preferred; rows scraped before it
    # existed fall back to "has a play count" so behaviour is unchanged for them.
    if "content_type" not in df.columns:
        df["content_type"] = ""
    df["content_type"] = df["content_type"].fillna("").astype(str)
    _IMG_CT = {"image", "graphimage", "photo", "sidecar", "graphsidecar", "carousel"}
    _VID_CT = {"video", "graphvideo", "reel", "clips", "igtv"}
    def _is_video_row(ct, plays) -> bool:
        c = str(ct or "").strip().lower()
        if c in _IMG_CT:
            return False
        if c in _VID_CT:
            return True
        return int(plays or 0) > 0
    df["_is_video"] = [
        _is_video_row(ct, pc) for ct, pc in zip(df["content_type"], df["play_count"])
    ]

    # ── Builder layout: which sheets/columns appear and in what order ───────────
    # The Exporter's "Advanced export settings" sends this. An empty/missing layout
    # means the full default workbook (all sheets, all columns) — so older callers
    # and the scheduled-report path are unaffected.
    _lay = layout if isinstance(layout, dict) else {}
    def _sub(key):
        v = _lay.get(key)
        return v if isinstance(v, dict) else {}
    _sum, _det, _nts = _sub("summary"), _sub("details"), _sub("notes")
    def _flag(d, k):
        v = d.get(k, True)
        return True if v is None else bool(v)
    sum_enabled = _flag(_sum, "enabled")
    det_enabled = _flag(_det, "enabled")
    nts_enabled = _flag(_nts, "enabled")
    sum_images  = _flag(_sum, "images")
    sum_dates   = _flag(_sum, "dates")
    sum_kpi     = _flag(_sum, "kpi")
    sum_videos  = _flag(_sum, "videos")
    det_type    = _flag(_det, "type")
    det_date    = _flag(_det, "date")
    det_range   = _flag(_det, "scrape_range")
    det_sort    = _flag(_det, "sort_order")
    det_url     = _flag(_det, "url")
    sheet_order = _lay.get("order") if isinstance(_lay.get("order"), list) else ["summary", "details", "notes"]

    # ── Calculated-metric columns for the Video Details sheet ───────────────────
    # Honour the metrics the user selected at scrape time. Only those derivable
    # from views/likes/comments/shares are computed here — VTR and CPV need a
    # separate view count / rate that profile audits don't capture.
    _CALC_FORMULAS = {
        "Engagement Rate":    lambda v, l, c, s: ((l + c + s) / v * 100) if v else 0.0,
        "Applause Rate":      lambda v, l, c, s: (l / v * 100) if v else 0.0,
        "Virality Rate":      lambda v, l, c, s: (s / v * 100) if v else 0.0,
        "Comment/View Ratio": lambda v, l, c, s: (c / v * 100) if v else 0.0,
    }
    # CPV ($) = rate / views per video. Computable only when per-KOL rates were
    # supplied at export time; otherwise it's listed as unavailable.
    rates_lower = {}
    for _k, _v in (rates or {}).items():
        try:
            rates_lower[str(_k).lower()] = float(_v)
        except (TypeError, ValueError):
            pass
    cpv_on = ("CPV ($)" in (calc_metrics or [])) and len(rates_lower) > 0

    sel_calc = [m for m in (calc_metrics or []) if m in _CALC_FORMULAS]
    if cpv_on:
        sel_calc = sel_calc + ["CPV ($)"]
    if not sel_calc:
        sel_calc = ["Engagement Rate"]   # sensible default so the sheet is never bare
    unsupported_calc = [m for m in (calc_metrics or [])
                        if m == "VTR" or (m == "CPV ($)" and not cpv_on)]

    # Optional raw engagement columns (default all when unspecified).
    _sel_raw = set(raw_metrics) if raw_metrics else {"Likes", "Comments", "Shares"}
    raw_cols = [rc for rc in ("Likes", "Comments", "Shares") if rc in _sel_raw]
    scrape_range = (f"{date_from or 'start'} -> {date_to or 'now'}"
                    if (date_from or date_to) else "All time")

    def _sort_group(g: pd.DataFrame) -> pd.DataFrame:
        if sort_by == "Most Views":   return g.sort_values("play_count", ascending=False)
        if sort_by == "Least Views":  return g.sort_values("play_count", ascending=True)
        # Date-based sorting: parse dates properly, push empty dates to the end
        if "post_date" in g.columns:
            g = g.copy()
            g["_date_sort"] = pd.to_datetime(g["post_date"], errors="coerce")
            # NaT (missing dates) go to the END regardless of sort direction
            if sort_by == "Most Recent":
                g = g.sort_values("_date_sort", ascending=False, na_position="last")
            else:  # "Oldest"
                g = g.sort_values("_date_sort", ascending=True, na_position="last")
            return g.drop("_date_sort", axis=1)
        # No date column → fall back to play_count
        return g.sort_values("play_count", ascending=(sort_by == "Oldest"))

    # Preserve paste order — use first-appearance of each username, not alphabetical
    ordered_kols = list(dict.fromkeys(df["username"].tolist()))
    # The V1, V2… columns show videos only, so width them off the video count.
    max_videos = max(
        (int(((df["username"] == k) & df["_is_video"]).sum()) for k in ordered_kols),
        default=0,
    )
    # Cap columns to exactly the requested number of videos (the DB can hold more
    # than one scrape's worth of rows per creator; we only show `limit`).
    if limit and limit > 0:
        max_videos = min(max_videos, limit)
    if not sum_videos:
        max_videos = 0   # builder hid the per-video V1, V2… columns

    # ── Colour palette ────────────────────────────────────────────────────────
    NAVY   = "1B3A6B"; WHITE = "FFFFFF"; LBLUE = "EBF3FB"
    GREEN  = "E8F8EE"; AMBER = "FEF3E2"
    VID_A  = "FFFFFF"; VID_B = "F0F4FF"   # alternating video columns
    thin   = Side(style="thin", color="BDC3C7")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfont(bold=True, color=WHITE, size=9):
        return Font(bold=bold, color=color, size=size, name="Calibri")
    def cell_font(bold=False, color="000000", size=9):
        return Font(bold=bold, color=color, size=size, name="Calibri")
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def align(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Summary — compact one-row-per-creator with cell comments
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = f"KOL Views ({platform_label})"

    fixed_headers = ["KOL / Creator", "Platform", "# Videos"]
    if sum_images: fixed_headers.append("# Images")
    fixed_headers += ["Avg Views", "Most Views", "Least Views"]
    if sum_dates:  fixed_headers += ["Date (Most Viewed)", "Date (Least Viewed)"]
    if sum_kpi:    fixed_headers.append("KPI Est. Views (next video)")
    opt_headers = []
    if incl_top5: opt_headers += ["Top 5 Avg Views"]
    if incl_bot5: opt_headers += ["Bottom 5 Avg Views"]
    video_headers = [f"V{i}" for i in range(1, max_videos+1)]
    all_headers = fixed_headers + opt_headers + video_headers

    # Write header row
    ws.append(all_headers)
    for cell in ws[1]:
        cell.fill      = fill(NAVY)
        cell.font      = hfont()
        cell.alignment = align()
        cell.border    = BORDER
        if cell.value == "KPI Est. Views (next video)":
            cell.comment = Comment(
                "Estimated views if you invest in this KOL again.\n\n"
                "Methodology: median of their historical video view counts, "
                "rounded to the nearest 10,000.\n\n"
                "Median is used instead of average so a single viral outlier "
                "doesn't inflate the expectation — this is meant to be a "
                "conservative, defensible number for budget planning.",
                "Total Scraper", width=280, height=140
            )
    ws.row_dimensions[1].height = 28

    # Write data rows
    for row_idx, kol in enumerate(ordered_kols, 2):
        group = df[df["username"] == kol]
        group   = _sort_group(group).reset_index(drop=True)
        # Videos drive every view-based number; images are counted but set aside.
        vid_group = group[group["_is_video"]].reset_index(drop=True)
        if limit and limit > 0:
            vid_group = vid_group.head(limit).reset_index(drop=True)   # exactly `limit`, no more
        n_images = int((~group["_is_video"]).sum())
        plays   = vid_group["play_count"]
        n       = len(vid_group)
        avg_v   = int(plays.mean()) if n else 0
        max_v   = int(plays.max())  if n else 0
        min_v   = int(plays.min())  if n else 0

        # Find dates for most and least viewed
        if n:
            max_row = vid_group.loc[plays.idxmax()]
            min_row = vid_group.loc[plays.idxmin()]
            date_most  = str(max_row.get("post_date","") or "")[:10]
            date_least = str(min_row.get("post_date","") or "")[:10]
        else:
            date_most = date_least = ""

        profile_url = (f"https://www.tiktok.com/@{kol}" if is_tiktok
                       else f"https://www.instagram.com/{kol}/")

        # KPI estimate: expected views from investing in this KOL, based on their
        # historical performance. Median (not mean) is used so a single viral outlier
        # doesn't inflate the expectation — this is meant to be a conservative,
        # defensible number for budget conversations. Rounded to the nearest 10,000.
        kpi_est = int(round(plays.median() / 10000.0)) * 10000 if n else 0

        row_data = [kol, platform_label, n]
        if sum_images: row_data.append(n_images)
        row_data += [avg_v, max_v, min_v]
        if sum_dates:  row_data += [date_most, date_least]
        if sum_kpi:    row_data.append(kpi_est)
        if incl_top5:
            top5 = vid_group.nlargest(5, "play_count")
            row_data.append(int(top5["play_count"].mean()) if len(top5) else 0)
        if incl_bot5:
            bot5 = vid_group.nsmallest(5, "play_count")
            row_data.append(int(bot5["play_count"].mean()) if len(bot5) else 0)

        if sum_videos:
            # Video cells: only the view count — URL and date go in the cell comment
            for i, (_, vrow) in enumerate(vid_group.iterrows()):
                row_data.append(int(vrow["play_count"]))
            # Pad to max_videos
            for _ in range(n, max_videos):
                row_data.append(None)

        ws.append(row_data)

        # Style the data row
        col_count = len(all_headers)
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row_idx, col_idx)
            header = all_headers[col_idx - 1]

            # Background
            if header == "KPI Est. Views (next video)":
                bg = "FFF4D6"  # gold highlight — this is the headline number for budget talks
            elif col_idx <= len(fixed_headers):
                bg = LBLUE
            elif "Top 5"    in header: bg = GREEN
            elif "Bottom 5" in header: bg = AMBER
            else:
                vid_num = col_idx - len(fixed_headers) - len(opt_headers)
                bg = VID_A if vid_num % 2 == 1 else VID_B

            cell.fill   = fill(bg)
            cell.border = BORDER
            cell.font   = cell_font(bold=(col_idx == 1))

            # Alignment
            if isinstance(cell.value, int):
                cell.alignment = align("center")
                cell.number_format = "#,##0"
            else:
                cell.alignment = align("left")

            # Cell comment for video columns: URL + date
            if header.startswith("V") and header[1:].isdigit():
                vid_idx = int(header[1:]) - 1
                if vid_idx < n:
                    vrow   = vid_group.iloc[vid_idx]
                    url    = str(vrow.get("post_url","") or "")
                    pdate  = str(vrow.get("post_date","") or "")[:10]
                    likes  = int(vrow.get("likes",0) or 0)
                    cmts   = int(vrow.get("comments",0) or 0)
                    caption_preview = str(vrow.get("caption","") or "")[:80]
                    comment_text = (
                        f"🔗 {url}\n"
                        f"📅 Posted: {pdate or 'unknown'}\n"
                        f"❤️ Likes: {likes:,}  💬 Comments: {cmts:,}"
                        + (f"\n📝 {caption_preview}..." if caption_preview else "")
                    )
                    comment = Comment(comment_text, "Total Scraper")
                    comment.width  = 280
                    comment.height = 100
                    cell.comment = comment

        ws.row_dimensions[row_idx].height = 18

    # Column widths for Sheet 1
    width_map = {
        "KOL / Creator": 26, "Platform": 12, "# Videos": 10, "# Images": 10,
        "Avg Views": 14, "Most Views": 14, "Least Views": 14,
        "KPI Est. Views (next video)": 18,
        "Date (Most Viewed)": 16, "Date (Least Viewed)": 16,
        "Top 5 Avg Views": 16, "Bottom 5 Avg Views": 16,
    }
    for col_idx, header in enumerate(all_headers, 1):
        w = width_map.get(header, 11 if header.startswith("V") else 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "B2"

    # Note row at top explaining comments (only meaningful when V-cells are shown)
    if sum_videos:
        ws.insert_rows(1)
        note_cell = ws.cell(1, 1,
            "💡 Hover over any V1, V2, V3… cell to see the video link, date posted, likes and comments.")
        note_cell.font      = Font(italic=True, color="5B6A8A", size=8, name="Calibri")
        note_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(all_headers), 12))
        ws.row_dimensions[1].height = 16

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Full Details — one row per video, all columns visible
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Video Details")
    detail_headers = ["KOL / Creator", "Platform"]
    if det_type: detail_headers.append("Type")
    detail_headers += ["Video #", "Views"]
    if det_date: detail_headers.append("Date Posted")
    detail_headers += raw_cols + sel_calc
    if det_range: detail_headers.append("Scrape Range")
    if det_sort:  detail_headers.append("Sort Order")
    if det_url:   detail_headers.append("Video URL")
    ws2.append(detail_headers)
    for cell in ws2[1]:
        cell.fill = fill(NAVY); cell.font = hfont(); cell.border = BORDER
        cell.alignment = align()
    ws2.row_dimensions[1].height = 24

    left_cols = {"KOL / Creator", "Platform", "Date Posted", "Scrape Range", "Sort Order", "Video URL"}
    detail_row = 2
    for kol in ordered_kols:
        group = df[df["username"] == kol]
        group = _sort_group(group).reset_index(drop=True)
        vid_group = group[group["_is_video"]]
        if limit and limit > 0:
            vid_group = vid_group.head(limit)                  # match the summary cap (videos)
        img_group = group[~group["_is_video"]]
        # Videos first (capped to the limit), then every image post so photos
        # always show up here even when the video limit is reached.
        details_group = pd.concat([vid_group, img_group]).reset_index(drop=True)
        for i, (_, vrow) in enumerate(details_group.iterrows(), 1):
            v = int(vrow["play_count"])
            is_vid = bool(vrow["_is_video"])
            l = int(vrow.get("likes", 0) or 0)
            c = int(vrow.get("comments", 0) or 0)
            s = int(vrow.get("shares", 0) or 0)
            row = [kol, platform_label]
            if det_type: row.append("Video" if is_vid else "Image")
            row += [i, v]
            if det_date: row.append(str(vrow.get("post_date", "") or "")[:10])
            _raw_vals = {"Likes": l, "Comments": c, "Shares": s}
            row += [_raw_vals[rc] for rc in raw_cols]
            for m in sel_calc:
                if v <= 0:
                    row.append("N/A")   # image / no-view post: view-based metric is undefined
                elif m == "CPV ($)":
                    _rate = rates_lower.get(str(kol).lower(), 0.0)
                    row.append(round(_rate / v, 4) if _rate else 0.0)
                else:
                    row.append(round(_CALC_FORMULAS[m](v, l, c, s), 2))
            if det_range: row.append(scrape_range)
            if det_sort:  row.append(sort_by)
            if det_url:   row.append(str(vrow.get("post_url", "") or ""))
            ws2.append(row)
            bg = LBLUE if i % 2 == 0 else WHITE
            for col_idx, cell in enumerate(ws2[detail_row], 1):
                cell.fill = fill(bg); cell.border = BORDER; cell.font = cell_font()
                hdr = detail_headers[col_idx - 1]
                cell.alignment = align("left" if hdr in left_cols else "center")
                if hdr == "CPV ($)":
                    cell.number_format = '"$"0.0000'
                elif hdr in sel_calc:
                    cell.number_format = '0.00"%"'
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
            ws2.row_dimensions[detail_row].height = 16
            detail_row += 1

    for col_idx, header in enumerate(detail_headers, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = (
            28 if header in ("KOL / Creator","Video URL") else
            16 if header == "Scrape Range" else
            14 if header in ("Date Posted","Sort Order") else 13
        )
    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Export Notes
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Export Notes")
    notes = [
        ("Total Scraper — Profile Feed (Audit) Export", True),
        (f"Platform: {platform_label}", False),
        (f"Video Sort Order: {sort_by}", False),
        (f"Scrape Date Range: {scrape_range}", False),
        (f"Top 5 included: {incl_top5}   |   Bottom 5 included: {incl_bot5}", False),
        ("", False),
        ("HOW TO READ THE KOL VIEWS SHEET", True),
        ("• Each V1, V2, V3… column = one video, sorted by your chosen order", False),
        ("• The cell value is the view count — select a range of V-cells to sum/average in the formula bar", False),
        ("• Hover over any V-cell to see the video link, date posted, likes and comments in the cell comment", False),
        ("", False),
        ("DISCLAIMERS", True),
        ("• View counts shown are from the Instagram platform only — not Facebook, Messenger, or Audience Network.", False),
        ("• 0 views means the creator's privacy settings prevented the API from returning this metric.", False),
        ("• Instagram counts a view when the video is watched for 3+ continuous seconds.", False),
        ("• TikTok Play Count increments every time the video starts, including auto-play and loops.", False),
        ("• Scraped metrics reflect values at time of collection and may differ from current on-app figures.", False),
    ]

    # ── Calculated metrics shown in the Video Details sheet ─────────────────────
    _calc_desc = {
        "Engagement Rate":    "(Likes + Comments + Shares) / Views x 100%",
        "Applause Rate":      "Likes / Views x 100%",
        "Virality Rate":      "Shares / Views x 100%",
        "Comment/View Ratio": "Comments / Views x 100%",
        "CPV ($)":            "Rate ($) / Views, per video",
    }
    notes += [("", False), ("CALCULATED METRICS (Video Details sheet)", True)]
    for m in sel_calc:
        notes.append((f"• {m}: {_calc_desc.get(m, '')}", False))
    if unsupported_calc:
        notes.append((f"• {', '.join(unsupported_calc)}: not shown — needs a separate view "
                      f"count / cost figure that a profile audit doesn't capture.", False))

    # ── Content types: reels carry views, photos don't ──────────────────────────
    notes += [
        ("", False),
        ("CONTENT TYPES (Video vs Image)", True),
        ("• Reels/videos carry a view (play) count; photos and carousels do not.", False),
        ("• The KOL Views sheet ranks and averages VIDEOS only — '# Images' shows how many "
         "photo posts were set aside so they don't drag the view averages to zero.", False),
        ("• In Video Details, image posts still list their Likes/Comments, but view-based metrics "
         "(Engagement Rate, Applause Rate, etc.) show 'N/A' — there is no view count to divide by.", False),
        ("• A video with 0 recorded views (private/restricted) is likewise shown as 'N/A' for "
         "view-based metrics.", False),
    ]

    # ── Data completeness — flag creators that returned nothing / fewer than asked ──
    try:
        present_lower = {str(u).lower() for u in df["username"].unique()}
        req = [str(u) for u in (requested_usernames or [])]
        missing = [u for u in req if u.lower() not in present_lower]
        partial = []
        if limit and limit > 0:
            for k in ordered_kols:
                # `limit` is a videos-per-creator target, so count videos, not photos.
                shown = min(int(((df["username"] == k) & df["_is_video"]).sum()), limit)
                if shown < limit:
                    partial.append((k, shown))
        comp = [("", False), ("DATA COMPLETENESS", True)]
        if limit and limit > 0:
            comp.append((f"Requested up to {limit} videos per creator.", False))
        if not missing and not partial:
            comp.append(("• All requested creators returned the full number of videos.", False))
        for u in missing:
            comp.append((f"• @{u}: NO DATA returned. Likely cause: private/restricted account, "
                         f"an Apify actor error, or a connectivity/rate-limit issue. Re-scrape to retry.", False))
        for k, cnt in partial:
            comp.append((f"• @{k}: {cnt}/{limit} videos (fewer than requested). Likely cause: a narrow date "
                         f"range, a low-posting creator, or partial rate-limiting.", False))
        notes = notes + comp
    except Exception:
        pass

    for text, bold in notes:
        ws3.append([text])
        ws3.cell(ws3.max_row, 1).font = Font(bold=bold, size=9, name="Calibri")
    ws3.column_dimensions["A"].width = 110

    # ── Apply the builder's sheet enable + order ────────────────────────────────
    # All three sheets are built above; here we drop the disabled ones and reorder
    # the rest. If the layout would leave nothing, the summary is kept so the file
    # is never empty (openpyxl can't save a workbook with zero sheets).
    _sheet_by_key = {"summary": ws, "details": ws2, "notes": ws3}
    _enabled = {"summary": sum_enabled, "details": det_enabled, "notes": nts_enabled}
    if not any(_enabled.values()):
        _enabled["summary"] = True
    for key, sheet in _sheet_by_key.items():
        if not _enabled[key]:
            wb.remove(sheet)
    desired = []
    for key in sheet_order:
        sheet = _sheet_by_key.get(key)
        if sheet is not None and _enabled.get(key) and sheet not in desired:
            desired.append(sheet)
    for sheet in list(wb.worksheets):          # append any not named in the order
        if sheet not in desired:
            desired.append(sheet)
    wb._sheets = desired
    wb.active = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
